#!/usr/bin/env python3
"""
sales_analysis.py — turn a sales source into a normalized analysis the brief is built from.

Usage:
    python sales_analysis.py <source> [days] [--metric revenue|velocity|blend] [--file PATH]

Sources:
    csv / xlsx     Read a sales export (the universal path — works with any business).
                   Pass the file with --file, or give the path directly as <source>.
    stripe paypal quickbooks plaid
                   Live connectors. These read their access token from the environment
                   (e.g. STRIPE_ACCESS_TOKEN, PAYPAL_*, QUICKBOOKS_*, PLAID_*). They are
                   thin wrappers around the same analysis core; if the connector library
                   or credentials are missing, the script exits 2 with a clear message so
                   the caller can fall back (see reference.md § Fallbacks).

Output: a single JSON object on stdout. Exit codes: 0 ok, 2 source/connector error.

The point of this script is determinism: rankings, period-over-period movement, and
data-sufficiency are computed here so the model never eyeballs numbers. The model turns
this JSON into the brief — it should not recompute the analysis by hand.
"""
import sys, os, json, argparse, datetime as dt
from collections import defaultdict

# ----- tunables (kept simple and explained so they can be reasoned about) -----
TREND_UP = 0.15      # >= +15% period-over-period = trending up
TREND_DOWN = -0.15   # <= -15% = trending down (a decliner worth flagging)
SLOW_REV_PCTILE = 0.20   # bottom 20% by revenue (and not trending up) = slow mover
SUFFICIENT_DAYS = 90     # < this much history => lean on benchmarks, label as directional


def die(msg):
    """Print to stderr and exit 2 — the caller's signal to fall back (try another
    connector, or proceed on partial/CSV data). See reference.md § Fallbacks."""
    sys.stderr.write(msg.rstrip() + "\n")
    sys.exit(2)


def _f(x, default=0.0):
    try:
        return float(str(x).replace("$", "").replace(",", "").replace("%", "").strip())
    except Exception:
        return default


def _read_rows(path):
    """Read csv or xlsx into a list of dict rows with normalized lowercase keys."""
    ext = os.path.splitext(path)[1].lower()
    if ext in (".xlsx", ".xls"):
        try:
            from openpyxl import load_workbook
        except ImportError:
            die("openpyxl not installed — export the sheet to CSV and pass that instead. [exit 2]")
        wb = load_workbook(path, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        header = [str(h).strip().lower() if h is not None else "" for h in rows[0]]
        return [dict(zip(header, r)) for r in rows[1:] if any(c is not None for c in r)]
    import csv
    with open(path, newline="", encoding="utf-8-sig") as fh:
        rdr = csv.DictReader(fh)
        return [{(k or "").strip().lower(): v for k, v in row.items()} for row in rdr]


def _find(keys, *cands):
    """Find the first column name containing any candidate substring."""
    for c in cands:
        for k in keys:
            if c in k:
                return k
    return None


def analyze_summary(rows, metric):
    """
    Summary shape: one row per product with a name, price, optional margin,
    and two period unit columns (prev, curr). This is the most common export.
    """
    keys = list(rows[0].keys())
    name_k = _find(keys, "product", "item", "name", "sku")
    price_k = _find(keys, "price", "unit price")
    margin_k = _find(keys, "margin")
    cat_k = _find(keys, "category", "type")
    # two period unit columns: anything with "unit" (or "qty"/"sold"); keep order
    unit_cols = [k for k in keys if ("unit" in k or "qty" in k or "sold" in k)]
    period_cols = unit_cols[-2:] if len(unit_cols) >= 2 else unit_cols
    prev_k = period_cols[0] if len(period_cols) == 2 else None
    curr_k = period_cols[-1] if period_cols else None

    products = []
    for r in rows:
        if not r.get(name_k):
            continue
        price = _f(r.get(price_k), 1.0)
        curr_u = _f(r.get(curr_k))
        prev_u = _f(r.get(prev_k)) if prev_k else 0.0
        rev = round(price * curr_u, 2)
        change = round((curr_u - prev_u) / prev_u, 4) if prev_u else None
        products.append({
            "name": str(r.get(name_k)).strip(),
            "category": (str(r.get(cat_k)).strip() if cat_k and r.get(cat_k) else None),
            "units": curr_u, "prev_units": prev_u, "price": price,
            "revenue": rev,
            "margin_pct": (_f(r.get(margin_k)) if margin_k else None),
            "pop_change": change,
        })
    return products, ("revenue" if metric != "velocity" else "velocity")


def analyze_transactions(rows, days, metric):
    """Transaction shape: many rows of date + product + amount (+ optional qty)."""
    keys = list(rows[0].keys())
    date_k = _find(keys, "date", "created", "time")
    name_k = _find(keys, "product", "item", "description", "name", "memo")
    amt_k = _find(keys, "amount", "total", "revenue", "price", "value")
    qty_k = _find(keys, "qty", "quantity", "units")
    agg = defaultdict(lambda: {"revenue": 0.0, "units": 0.0})
    dates = []
    for r in rows:
        name = str(r.get(name_k) or "Uncategorized sale").strip()
        agg[name]["revenue"] += _f(r.get(amt_k))
        agg[name]["units"] += _f(r.get(qty_k), 1.0)
        d = str(r.get(date_k) or "")[:10]
        try:
            dates.append(dt.date.fromisoformat(d))
        except Exception:
            pass
    span = (max(dates) - min(dates)).days + 1 if dates else (days or 0)
    weeks = max(span / 7.0, 1.0)
    products = []
    for name, v in agg.items():
        products.append({
            "name": name, "category": None,
            "units": round(v["units"], 2), "prev_units": None,
            "price": round(v["revenue"] / v["units"], 2) if v["units"] else 0.0,
            "revenue": round(v["revenue"], 2), "margin_pct": None,
            "velocity_per_week": round(v["units"] / weeks, 2),
            "pop_change": None,
        })
    return products, span


def build_output(products, metric, sort_key, days_of_data, granularity, unavailable=None):
    rank_key = {
        "revenue": lambda p: p.get("revenue") or 0,
        "velocity": lambda p: p.get("velocity_per_week") or p.get("units") or 0,
        "blend": lambda p: ((p.get("revenue") or 0) ** 0.5) * ((p.get("velocity_per_week") or p.get("units") or 0) + 1),
    }.get(metric, lambda p: p.get("revenue") or 0)
    ranked = sorted(products, key=rank_key, reverse=True)

    revs = sorted([p["revenue"] for p in products])
    cut = revs[int(len(revs) * SLOW_REV_PCTILE)] if revs else 0
    up = [p for p in products if p.get("pop_change") is not None and p["pop_change"] >= TREND_UP]
    down = [p for p in products if p.get("pop_change") is not None and p["pop_change"] <= TREND_DOWN]
    slow = [p for p in products if p["revenue"] <= cut and p not in up]

    def slim(p):
        return {k: p[k] for k in ("name", "category", "units", "revenue", "margin_pct", "pop_change", "velocity_per_week") if k in p and p[k] is not None}

    out = {
        "data_sufficiency": {
            "sufficient": days_of_data is None or days_of_data >= SUFFICIENT_DAYS,
            "days_of_data": days_of_data,
            "recommend_benchmarks": bool(days_of_data is not None and days_of_data < SUFFICIENT_DAYS),
        },
        "granularity": granularity,
        "metric": metric,
        "requested_metric_unavailable": unavailable,
        "totals": {
            "revenue": round(sum(p["revenue"] for p in products), 2),
            "products": len(products),
        },
        "top_performers": [slim(p) for p in ranked[:5]],
        "trending_up": sorted([slim(p) for p in up], key=lambda x: x.get("pop_change", 0), reverse=True),
        "trending_down": sorted([slim(p) for p in down], key=lambda x: x.get("pop_change", 0)),
        "slow_movers": [slim(p) for p in slow][:5],
    }
    return out


def from_file(path, days, metric):
    rows = _read_rows(path)
    if not rows:
        die("No rows found in the file. [exit 2]")
    keys = list(rows[0].keys())
    unavailable = "margin" if metric == "margin" else None
    if unavailable:
        metric = "revenue"  # margin needs unit costs the feeds don't carry; fall back
    is_txn = _find(keys, "date", "created", "time") and _find(keys, "amount", "total", "value")
    if is_txn:
        products, span = analyze_transactions(rows, days, metric)
        return build_output(products, metric, "revenue", span, "product", unavailable)
    products, _ = analyze_summary(rows, metric)
    return build_output(products, metric, "revenue", days, "product", unavailable)


def from_connector(source, days, metric):
    """
    Live connectors read a token from the environment and pull recent sales, then run
    the same analysis core. Kept thin on purpose. If creds/libs are missing we exit 2
    with an actionable message rather than pretending — the caller falls back per
    reference.md (try another connector, or proceed on partial/CSV data).
    """
    token_env = {
        "stripe": "STRIPE_ACCESS_TOKEN", "paypal": "PAYPAL_ACCESS_TOKEN",
        "quickbooks": "QUICKBOOKS_ACCESS_TOKEN", "plaid": "PLAID_ACCESS_TOKEN",
    }[source]
    if not os.environ.get(token_env):
        die(f"{source}: no {token_env} in environment — connect {source} or use a CSV export. [exit 2]")
    # Live API extraction plugs in here (each source normalizes into the same product list,
    # then calls build_output). Plaid sets granularity 'revenue_source'; others 'product'.
    die(f"{source}: live extraction not configured in this environment — export to CSV and pass --file. [exit 2]")


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("source")
    ap.add_argument("days", nargs="?", type=int, default=90)
    ap.add_argument("--metric", default="revenue", choices=["revenue", "velocity", "blend", "margin"])
    ap.add_argument("--file")
    a = ap.parse_args()

    path = a.file
    if a.source in ("csv", "xlsx", "file") or (path is None and os.path.exists(a.source)):
        path = path or a.source
    if path:
        out = from_file(path, a.days, a.metric)
    elif a.source in ("stripe", "paypal", "quickbooks", "plaid"):
        out = from_connector(a.source, a.days, a.metric)
    else:
        die(f"Unknown source '{a.source}'. Use a connector name or pass --file path.csv [exit 2]")
    json.dump(out, sys.stdout, indent=2)
    print()


if __name__ == "__main__":
    main()
